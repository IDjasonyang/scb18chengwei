# 实战测试


import requests
import openpyxl
# 读取测试用例
def read_case(filename,sheetname):
    wb = openpyxl.load_workbook(filename)  #加载工作簿，打开一个excel文件
    sheet = wb[sheetname]  #打开某一个表单
    row_max = sheet.max_row  #获取最大行数
    case_list = []   #新建空列表，存放for循环依次读取到的测试用例数据
    for i in range(2,row_max+1):
        data_dict = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,  #读取URL值
        data = sheet.cell(row=i,column=6).value,  #读取data值
        expect = sheet.cell(row=i,column=7).value    #读取期望
        )
        case_list.append(data_dict)
    return case_list  #把每一行读取到的测试用例数据生成的字典，追加到List中


# 测试用例执行
def api_fun(url,data):
    headers = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}  #请求头
    res = requests.post(url=url,json=data,headers=headers).json()
    return res


# 写入测试结果
def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)   #加载工作簿，打开一个excel文件
    sheet = wb[sheetname]  #打开某一个表单
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)


#断言
def execute_fun(filename,sheetname):
    cases = read_case(filename,sheetname)    #调用函数,设置变量接收返回值
    for case in cases:
        case_id = case['case_id']   #获取用例编号
        url_case = case['url']      #获取url的值
        data = eval(case['data'])   #字符串--字典
        expect = eval(case['expect'])  #预期结果
        expect_msg = expect['msg']  #预期结果中的msg
        real_result = api_fun(url=url_case,data=data)
        real_msg = real_result['msg']  #获取实际结果中的msg
        print('期望结果为: {}'.format(expect_msg))
        print('实际结果为: {}'.format(real_msg))

        if expect_msg == real_msg:
            print('第{}条用例不通过'.format(case_id))
            final_re = 'Passed'
        else:
            print('第{}条用例不通过'.format(case_id))
            final_re = 'Failed'
        write_result(filename,sheetname,case_id+1,8,final_re)  #写入结果到测试用例
        print('**********')
#调用函数
execute_fun('../test_data/test_case_api.xlsx','register')
execute_fun('../test_data/test_case_api.xlsx','login')


